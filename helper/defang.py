# Defanging utilities for Kanvas: defang indicators (IPs, URLs, domains, emails) in text
# and Excel files to reduce accidental clicks; used for reports and export.
# Reviewed on 01/02/2026 by Jinto Antony

import logging
import re

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

IP_REGEX = re.compile(r"(\d{1,3}\.\d{1,3}\.\d{1,3})\.(\d{1,3})")
HTTP_REGEX = re.compile(r"(https?)(://)", re.IGNORECASE)
DOMAIN_REGEX = re.compile(
    r"([a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z0-9][-a-zA-Z0-9]*)\.([a-zA-Z]{2,})"
)


def defang_text(text):
    if not text or not isinstance(text, str):
        return text
    result = IP_REGEX.sub(r"\1[.]\2", text)
    result = HTTP_REGEX.sub(r"hxxp\2", result)
    result = DOMAIN_REGEX.sub(r"\1[.]\2", result)
    result = result.replace("@", "[at]")
    return result


def defang_excel_file(input_file_path, output_file_path, progress_callback=None):
    try:
        logger.info(
            "Starting defanging process for file: %s", input_file_path
        )
        workbook = load_workbook(input_file_path)
        total_sheets = len(workbook.sheetnames)
        logger.info("Processing %d sheets", total_sheets)
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            if progress_callback:
                progress_callback(
                    sheet_idx, total_sheets, f"Defanging: {sheet_name}"
                )
            logger.info("Processing sheet: %s", sheet_name)
            sheet = workbook[sheet_name]
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        cell.value = defang_text(original_value)
                        if original_value != cell.value:
                            logger.debug(
                                "Defanged: '%s' -> '%s'",
                                original_value,
                                cell.value,
                            )
        workbook.save(output_file_path)
        logger.info("Defanged file saved to: %s", output_file_path)
        return True
    except Exception as e:
        logger.error("Error during defanging: %s", e)
        raise


def defang_string(text):
    return defang_text(text)
