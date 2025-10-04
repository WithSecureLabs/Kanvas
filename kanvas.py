import sys
import os
import sqlite3
import shutil
import re
import traceback
import logging
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
from PySide6 import __version__
from PySide6.QtWidgets import (
    QApplication, QMessageBox, QPushButton, QMainWindow, QFileDialog, QTreeView, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, 
    QComboBox, QGridLayout, QTextEdit, QSizePolicy, QWidget, QDateEdit, QProgressBar, QScrollArea, QHeaderView, QSplashScreen,
    QMenu, QStyledItemDelegate, QStyle, QToolTip
)
from PySide6.QtGui import QStandardItemModel, QStandardItem, QColor, QFont, QPixmap, QKeySequence, QAction, QPainter, QTextOption
from PySide6.QtCore import QFile, Qt, QDate, QRect, QTimer, QSize, QModelIndex
from viz_network import visualize_network
from viz_timeline import open_timeline_window
from helper.database_utils import create_all_tables
from helper.download_updates import download_updates
from helper.api_config import open_api_settings
from helper.mapping_defend import open_defend_window
from helper.mapping_attack import mitre_mapping
from helper.mapping_veris import open_veris_window
from helper.bookmarks import display_bookmarks_kb
from helper.lookup_entraid import open_entra_lookup_window
from helper.resources_data import display_msportals_data, display_event_id_kb
from helper.lookup_domain import open_domain_lookup_window
from helper.lookup_cve import open_cve_window
from helper.lookup_ip import open_ip_lookup_window
from helper.lookup_file import open_hash_lookup_window
from helper.system_type import EvidenceTypeManager
from helper.lookup_email import open_email_lookup_window
from helper.lookup_ransomware import open_ransomware_kb_window
from helper.lolbas import display_lolbas_kb
from helper.stix import convert_indicators_to_stix
from helper.defang import defang_excel_file
from markdown_editor import handle_markdown_editor
from helper.bot import handle_chat_bot
from helper.mitre_flow_platform import open_mitre_flow_window
from filelock import FileLock, Timeout
from PySide6.QtGui import QIcon
from shiboken6 import isValid 
from helper.windowsui import Ui_KanvasMainWindow
from helper.system_type import SystemTypeManager
from PySide6.QtGui import QFontDatabase

class CustomTreeItemDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.highlight_color = QColor(66, 139, 202, 50) 
        self.important_color = QColor(220, 53, 69, 50)   
    def paint(self, painter, option, index):
        text = index.data(Qt.DisplayRole) or ""
        painter.save()
        rect = option.rect
        painter.setRenderHint(QPainter.Antialiasing)
        if option.state & QStyle.State_Selected:
            painter.fillRect(rect, option.palette.highlight())
        elif option.state & QStyle.State_MouseOver:
            painter.fillRect(rect, QColor(227, 242, 253, 100))  
        elif index.row() % 2 == 1:  
            painter.fillRect(rect, QColor(248, 249, 250, 100))  
        text_rect = rect.adjusted(8, 4, -8, -4)
        important_keywords = ['critical', 'high', 'alert', 'error', 'failed', 'blocked']
        is_important = any(keyword.lower() in text.lower() for keyword in important_keywords)
        
        if is_important:
            painter.setPen(QColor(220, 53, 69))  
            font = painter.font()
            font.setBold(True)
            painter.setFont(font)
        painter.drawText(text_rect, Qt.AlignLeft | Qt.AlignVCenter, text)
        painter.restore()
    def sizeHint(self, option, index):
        text = index.data(Qt.DisplayRole) or ""
        if not text:
            return QSize(100, 24)
        font_metrics = option.fontMetrics
        text_rect = font_metrics.boundingRect(text)
        width = max(100, text_rect.width() + 16)
        height = max(24, text_rect.height() + 8)
        return QSize(width, height)
    def helpEvent(self, event, view, option, index):
        if event.type() == event.Type.ToolTip:
            text = index.data(Qt.DisplayRole) or ""
            if text and len(text) > 50: 
                QToolTip.showText(event.globalPos(), text, view)
                return True
        return super().helpEvent(event, view, option, index)
class MainApp:
    def __init__(self):
        self.app = QApplication(sys.argv)
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', filename='kanvas.log')
        self.logger = logging.getLogger(__name__)
        image_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images")
        logo_path = os.path.join(image_dir, "logo.png")
        app_icon = QIcon(logo_path)
        self.app.setWindowIcon(app_icon)
        splash_pixmap = QPixmap(logo_path)
        if not splash_pixmap.isNull():
            splash_pixmap = splash_pixmap.scaled(300, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.splash = QSplashScreen(splash_pixmap)
        else:
            splash_pixmap = QPixmap(300, 150)
            splash_pixmap.fill(QColor(40, 44, 52))
            self.splash = QSplashScreen(splash_pixmap)
            self.splash.showMessage("KANVAS\n \nLoading...", Qt.AlignCenter, QColor(255, 255, 255))
        self.splash.show()
        self.app.processEvents()
        self.child_windows = []
        self.mitre_flow_window = None
        self.db_path = "kanvas.db"
        self.file_lock = None 
        self.read_only_mode = False  
        self.splash.showMessage("Initializing database...", Qt.AlignBottom | Qt.AlignCenter, QColor(255, 255, 255))
        self.app.processEvents()
        create_all_tables(self.db_path)
        self.system_type_manager = SystemTypeManager(self.db_path)
        self.splash.showMessage("Loading user interface...", Qt.AlignBottom | Qt.AlignCenter, QColor(255, 255, 255))
        self.app.processEvents()
        self.window = self.load_ui()
        self.window.closeEvent = self.closeEvent
        self.evidence_type_manager = EvidenceTypeManager(self.db_path, self.window)
        self.current_workbook = None
        self.current_file_path = None
        self.current_sheet_name = None
        self.splash.showMessage("Connecting UI elements...", Qt.AlignBottom | Qt.AlignCenter, QColor(255, 255, 255))
        self.app.processEvents()
        self.connect_ui_elements()
        QTimer.singleShot(1000, self.finish_loading)
        QTimer.singleShot(2000, self.preload_mitre_flow)

    def get_monospace_font(self):
        if sys.platform == "darwin": 
            font_families = ["SF Mono", "Monaco", "Menlo", "Courier"]
        elif sys.platform == "win32":  
            font_families = ["Consolas", "Courier New", "Lucida Console"]
        else:  
            font_families = ["DejaVu Sans Mono", "Liberation Mono", "Courier"]
        for font in font_families:
            if font in QFontDatabase.families():
                return font
        return "Courier"  
    def get_sans_serif_font(self):
        if sys.platform == "darwin": 
            font_families = ["SF Pro Display", "Helvetica Neue", "Helvetica", "Arial"]
        elif sys.platform == "win32": 
            font_families = ["Segoe UI", "Arial", "Calibri", "Tahoma"]
        else: 
            font_families = ["Ubuntu", "DejaVu Sans", "Liberation Sans", "Arial"]
        for font in font_families:
            if font in QFontDatabase.families():
                return font
        return "Arial" 
    def finish_loading(self):
        self.window.showMaximized()
        self.splash.finish(self.window)
    
    def preload_mitre_flow(self):
        try:
            self.logger.info("Starting MITRE Flow background preloading")
            window = open_mitre_flow_window(self.window)
            if window:
                self.mitre_flow_window = window
                window.hide()
                self.logger.info("MITRE Flow window preloaded successfully")
            else:
                self.logger.warning("MITRE Flow window preloading failed")
        except Exception as e:
            self.logger.error(f"Error preloading MITRE Flow window: {str(e)}")
    def get_lock_path(self, excel_path):
        return f"{excel_path}.lock"
    def acquire_file_lock(self, file_path):
        try:
            if self.file_lock:
                self.release_file_lock()  
            lock_path = self.get_lock_path(file_path)
            self.file_lock = FileLock(lock_path, timeout=1) 
            try:
                self.file_lock.acquire(timeout=1)
                self.read_only_mode = False
                return True
            except Timeout:
                self.logger.info(f"File {file_path} is locked by another process")
                result = QMessageBox.question( self.window, "File is in use", f"The file is currently being edited by another user.\nDo you want to open it in read-only mode?", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                if result == QMessageBox.Yes:
                    self.read_only_mode = True
                    return False
                else:
                    return None  
        except Exception as e:
            self.logger.error(f"Error acquiring lock: {e}")
            self.file_lock = None
            return None  
    
    def release_file_lock(self):
        if self.file_lock:
            try:
                if hasattr(self.file_lock, 'is_locked') and self.file_lock.is_locked:
                    self.file_lock.release()
                    self.logger.info("File lock released")
            except Exception as e:
                self.logger.error(f"Error releasing file lock: {e}")
            finally:
                self.file_lock = None
    def closeEvent(self, event):
        self.release_file_lock()
        if self.mitre_flow_window:
            try:
                self.mitre_flow_window.close()
                self.mitre_flow_window = None
            except Exception as e:
                self.logger.error(f"Error closing preloaded MITRE Flow window: {e}")
        for window in self.child_windows:
            if window and hasattr(window, 'setEnabled'):
                window.setEnabled(False)
        windows_to_close = self.child_windows.copy()  
        for i, window in enumerate(windows_to_close):
            try:
                if window and hasattr(window, 'close'):
                    window.close()
                    QApplication.processEvents()  
            except Exception as e:
                self.logger.error(f"Error during first close attempt: {e}")
        remaining = [w for w in self.child_windows if hasattr(w, 'isVisible') and w.isVisible()]
        if remaining:
            self.logger.info(f"Found {len(remaining)} windows still open, forcing deletion...")
            for window in remaining:
                try:
                    self.logger.info(f"Force closing: {window.windowTitle() if hasattr(window, 'windowTitle') else 'Unknown'}")
                    window.setAttribute(Qt.WA_DeleteOnClose, True) 
                    window.setParent(None)  
                    window.close()
                    window.deleteLater()
                    QApplication.processEvents() 
                except Exception as e:
                    self.logger.error(f"Error during forced close: {e}")
        self.child_windows.clear()
        event.accept()

    def load_ui(self):
        window = QMainWindow()
        ui = Ui_KanvasMainWindow()
        ui.setupUi(window)
        window.ui = ui
        return window
    def connect_ui_elements(self):
        self.tree_view = self.window.ui.treeViewMain
        if self.tree_view:
            self.apply_treeview_styling()
            self.configure_treeview_properties()
            self.tree_view.doubleClicked.connect(self.edit_row)
            self.tree_view.setContextMenuPolicy(Qt.CustomContextMenu)
            self.tree_view.customContextMenuRequested.connect(self.show_context_menu)
        else:
            self.logger.warning("treeViewMain not found!")
        self.connect_button("left_button_2", self.handle_veris_window)
        self.connect_button("left_button_3", self.handle_defend_mapping)  
        self.connect_button("left_button_4", self.handle_mitre_mapping)
        self.connect_button("left_button_5", self.handle_visualize_network)
        self.connect_button("left_button_6", self.handle_timeline_window)
        self.connect_button("left_button_7", self.open_new_case_window)
        self.connect_button("left_button_8", self.load_data_into_treeview)
        self.connect_button("left_button_9", self.handle_ip_lookup)  
        self.connect_button("left_button_10", self.handle_ransomware_kb)
        self.connect_button("left_button_11", self.handle_domain_lookup) 
        self.connect_button("left_button_12", self.handle_hash_lookup) 
        self.connect_button("left_button_15", self.handle_markdown_editor)
        self.connect_button("left_button_16", self.display_bookmarks_kb)
        self.connect_button("left_button_18", self.open_api_settings)
        self.connect_button("left_button_19", self.handle_download_updates)
        self.connect_button("left_button_17", self.display_event_id_kb)
        self.connect_button("left_button_13", self.entra_appid)
        self.connect_button("left_button_14", self.handle_cve_lookup)
        self.connect_button("left_button_20", self.handle_chat_bot)
        self.connect_button("left_button_21", self.handle_mitre_flow)
        self.connect_button("left_button_22", self.handle_email_lookup)
        self.connect_button("down_button_1", self.add_new_row)  
        self.connect_button("down_button_2", self.delete_row)   
        self.connect_button("down_button_3", self.load_sheet)   
        self.connect_button("down_button_6", self.defang)     
        self.connect_button("down_button_7", self.handle_stix_export)
        self.connect_button("down_button_8", self.evidence_type_manager.show_add_evidence_type_dialog)
        self.connect_button("down_button_9", self.handle_add_system_type)
        self.connect_button("more_button", self.handle_more_button)
        self.setup_more_button_menu()
        self.hide_bottom_buttons()
        
    def connect_button(self, button_name, handler):
        button = getattr(self.window.ui, button_name, None)
        if button:
            button.clicked.connect(handler)
        else:
            self.logger.warning(f"{button_name} not found!")
    
    def hide_bottom_buttons(self):
        try:
            self.window.ui.footerLayout.removeWidget(self.window.ui.down_button_2)
            self.window.ui.footerLayout.removeWidget(self.window.ui.down_button_3)
            self.window.ui.down_button_2.hide()
            self.window.ui.down_button_3.hide()
            self.logger.info("Bottom buttons 'Delete Entry' and 'Refresh Table' have been hidden")
        except Exception as e:
            self.logger.error(f"Error hiding bottom buttons: {e}")
    
    def restore_bottom_buttons(self):
        try:
            self.window.ui.down_button_2.show()
            self.window.ui.down_button_3.show()
            self.window.ui.footerLayout.insertWidget(1, self.window.ui.down_button_2)
            self.window.ui.footerLayout.insertWidget(2, self.window.ui.down_button_3)
            self.logger.info("Bottom buttons 'Delete Entry' and 'Refresh Table' have been restored")
        except Exception as e:
            self.logger.error(f"Error restoring bottom buttons: {e}")
    
    def get_platform_font_settings(self):
        if sys.platform == "darwin":  
            return "font-size: 11pt; font-family: 'Helvetica Neue', 'Helvetica', Arial, sans-serif;"
        elif sys.platform == "win32":  
            return "font-size: 9pt; font-family: 'Segoe UI', 'Tahoma', Arial, sans-serif;"
        else:  
            return "font-size: 10pt; font-family: 'Ubuntu', 'DejaVu Sans', Arial, sans-serif;"
    def apply_treeview_styling(self):
        font_settings = self.get_platform_font_settings()
        modern_style = """
        QTreeView {{
            background-color: #ffffff;
            alternate-background-color: #f8f9fa;
            selection-background-color: #007acc;
            selection-color: white;
            border: 1px solid #dee2e6;
            border-radius: 8px;
            outline: none;
            gridline-color: #e9ecef;
            {}
        }}

        QTreeView::item {{
            padding: 8px 12px;
            border-bottom: 1px solid #f1f3f4;
            min-height: 24px;
        }}

        QTreeView::item:hover {{
            background-color: #e3f2fd;
            color: #1976d2;
        }}

        QTreeView::item:selected {{
            background-color: #007acc;
            color: white;
        }}

        QTreeView::item:selected:hover {{
            background-color: #005a9e;
        }}

        QTreeView::branch {{
            background: transparent;
        }}

        QTreeView::branch:has-children:!has-siblings:closed,
        QTreeView::branch:closed:has-children:has-siblings {{
            image: none;
            border-left: 5px solid transparent;
            border-right: 5px solid transparent;
            border-top: 5px solid #666666;
            width: 0px;
            height: 0px;
        }}

        QTreeView::branch:open:has-children:!has-siblings,
        QTreeView::branch:open:has-children:has-siblings {{
            image: none;
            border-left: 5px solid transparent;
            border-right: 5px solid transparent;
            border-bottom: 5px solid #666666;
            width: 0px;
            height: 0px;
        }}

        QHeaderView::section {{
            background-color: #f8f9fa;
            color: #495057;
            padding: 12px 16px;
            border: none;
            border-bottom: 2px solid #dee2e6;
            border-right: 1px solid #e9ecef;
            font-weight: bold;
            font-size: 10pt;
        }}

        QHeaderView::section:hover {{
            background-color: #e9ecef;
        }}

        QHeaderView::section:pressed {{
            background-color: #dee2e6;
        }}

        QScrollBar:vertical {{
            background-color: #f8f9fa;
            width: 12px;
            border-radius: 6px;
        }}

        QScrollBar::handle:vertical {{
            background-color: #cbd5e0;
            border-radius: 6px;
            min-height: 20px;
        }}

        QScrollBar::handle:vertical:hover {{
            background-color: #a0aec0;
        }}

        QScrollBar::add-line:vertical,
        QScrollBar::sub-line:vertical {{
            height: 0px;
        }}

        QScrollBar:horizontal {{
            background-color: #f8f9fa;
            height: 12px;
            border-radius: 6px;
        }}

        QScrollBar::handle:horizontal {{
            background-color: #cbd5e0;
            border-radius: 6px;
            min-width: 20px;
        }}

        QScrollBar::handle:horizontal:hover {{
            background-color: #a0aec0;
        }}

        QScrollBar::add-line:horizontal,
        QScrollBar::sub-line:horizontal {{
            width: 0px;
        }}
        """.format(font_settings)
        self.tree_view.setStyleSheet(modern_style)
    
    def configure_treeview_properties(self):
        custom_delegate = CustomTreeItemDelegate(self.tree_view)
        self.tree_view.setItemDelegate(custom_delegate)
        self.tree_view.setAlternatingRowColors(True)
        self.tree_view.setRootIsDecorated(True)
        self.tree_view.setSortingEnabled(True)
        self.tree_view.setSelectionBehavior(QTreeView.SelectRows)
        self.tree_view.setSelectionMode(QTreeView.ExtendedSelection)
        self.tree_view.setDragDropMode(QTreeView.NoDragDrop)
        self.tree_view.setDefaultDropAction(Qt.IgnoreAction)
        self.tree_view.setWordWrap(True)
        self.tree_view.setUniformRowHeights(True)
        self.tree_view.setToolTip("Click on items to view details. Use Ctrl+Click for multiple selections.")
        header = self.tree_view.header()
        if header:
            header.setStretchLastSection(True)
            header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            header.setHighlightSections(True)
            header.setSortIndicatorShown(True)
            header.setToolTip("Click column headers to sort. Drag to resize columns.")
        self.tree_view.resizeColumnToContents(0)
    
    def apply_standard_treeview_styling(self, tree_view):
        if not tree_view:
            return
        if sys.platform == "darwin":  
            font_settings = "font-size: 10pt; font-family: 'Helvetica Neue', 'Helvetica', Arial, sans-serif;"
        elif sys.platform == "win32":  
            font_settings = "font-size: 8pt; font-family: 'Segoe UI', 'Tahoma', Arial, sans-serif;"
        else:  
            font_settings = "font-size: 9pt; font-family: 'Ubuntu', 'DejaVu Sans', Arial, sans-serif;"
            
        standard_style = """
        QTreeView {{
            background-color: #ffffff;
            alternate-background-color: #f8f9fa;
            selection-background-color: #007acc;
            selection-color: white;
            border: 1px solid #dee2e6;
            border-radius: 8px;
            gridline-color: #e9ecef;
            {}
        }}
        QTreeView::item {{
            padding: 6px 10px;
            border-bottom: 1px solid #f1f3f4;
            min-height: 22px;
        }}
        QTreeView::item:hover {{
            background-color: #e3f2fd;
            color: #1976d2;
        }}
        QTreeView::item:selected {{
            background-color: #007acc;
            color: white;
        }}
        QTreeView::item:selected:hover {{
            background-color: #005a9e;
        }}
        QHeaderView::section {{
            background-color: #f8f9fa;
            color: #495057;
            padding: 10px 12px;
            border: none;
            border-bottom: 2px solid #dee2e6;
            border-right: 1px solid #e9ecef;
            font-weight: bold;
            font-size: 9pt;
        }}
        QHeaderView::section:hover {{
            background-color: #e9ecef;
        }}
        QHeaderView::section:pressed {{
            background-color: #dee2e6;
        }}
        QScrollBar:vertical {{
            background-color: #f8f9fa;
            width: 10px;
            border-radius: 5px;
        }}
        QScrollBar::handle:vertical {{
            background-color: #cbd5e0;
            border-radius: 5px;
            min-height: 20px;
        }}
        QScrollBar::handle:vertical:hover {{
            background-color: #a0aec0;
        }}
        """.format(font_settings)
        tree_view.setStyleSheet(standard_style)
        tree_view.setAlternatingRowColors(True)
        tree_view.setSelectionBehavior(QTreeView.SelectRows)
        tree_view.setSortingEnabled(True)
        tree_view.setWordWrap(True)
        tree_view.setUniformRowHeights(True)
        header = tree_view.header()
        if header:
            header.setStretchLastSection(True)
            header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            header.setHighlightSections(True)
            header.setSortIndicatorShown(True)
        tree_view.resizeColumnToContents(0)
    
    def show_context_menu(self, position):
        if not self.tree_view or not self.tree_view.model():
            return
        selected_indices = self.tree_view.selectionModel().selectedRows()
        has_selection = len(selected_indices) > 0
        context_menu = QMenu(self.tree_view)
        edit_action = QAction("✏️ Edit Row", self.tree_view)
        edit_action.triggered.connect(self.edit_row_from_context)
        edit_action.setEnabled(has_selection and not self.read_only_mode)
        context_menu.addAction(edit_action)
        add_action = QAction("➕ Add New Row", self.tree_view)
        add_action.triggered.connect(self.add_new_row)
        add_action.setEnabled(not self.read_only_mode)
        context_menu.addAction(add_action)
        context_menu.addSeparator()
        duplicate_action = QAction("📋 Duplicate Row", self.tree_view)
        duplicate_action.triggered.connect(self.duplicate_row)
        duplicate_action.setEnabled(has_selection and not self.read_only_mode)
        context_menu.addAction(duplicate_action)
        copy_action = QAction("📄 Copy Row Data", self.tree_view)
        copy_action.triggered.connect(self.copy_row_data)
        copy_action.setEnabled(has_selection)
        context_menu.addAction(copy_action)
        context_menu.addSeparator()
        delete_action = QAction("🗑️ Delete Row(s)", self.tree_view)
        delete_action.setShortcut(QKeySequence("Delete"))
        delete_action.triggered.connect(self.delete_row)
        delete_action.setEnabled(has_selection and not self.read_only_mode)
        context_menu.addAction(delete_action)
        context_menu.addSeparator()
        refresh_action = QAction("🔄 Refresh Data", self.tree_view)
        refresh_action.triggered.connect(self.load_sheet)
        context_menu.addAction(refresh_action)
        context_menu.addSeparator()
        analysis_menu = context_menu.addMenu("🔍 Analysis")
        network_analysis_action = QAction("Network Visualization", self.tree_view)
        network_analysis_action.triggered.connect(self.handle_visualize_network)
        analysis_menu.addAction(network_analysis_action)
        timeline_analysis_action = QAction("Timeline Analysis", self.tree_view)
        timeline_analysis_action.triggered.connect(self.handle_timeline_window)
        analysis_menu.addAction(timeline_analysis_action)
        mitre_analysis_action = QAction("MITRE Mapping", self.tree_view)
        mitre_analysis_action.triggered.connect(self.handle_mitre_mapping)
        analysis_menu.addAction(mitre_analysis_action)
        context_menu.exec(self.tree_view.mapToGlobal(position))
            
    def check_excel_loaded(self):
        if not self.current_workbook or not self.current_file_path:
            QMessageBox.warning(self.window, "Warning", "No Excel file loaded. Please load a file first.")
            return False
        return True

    def handle_veris_window(self):
        if self.check_excel_loaded():
            window = open_veris_window(self.window)
            self.track_child_window(window)

    def handle_mitre_mapping(self):
        if self.check_excel_loaded():
            window = mitre_mapping(self.window)
            self.track_child_window(window)

    def handle_visualize_network(self):
        if self.check_excel_loaded():
            window = visualize_network(self.window, self.system_type_manager)
            self.track_child_window(window)

    def handle_timeline_window(self):
        if self.check_excel_loaded():
            window = open_timeline_window(self.window)
            self.track_child_window(window)
            
    def handle_download_updates(self):
        window = download_updates(self.window)
        self.track_child_window(window)
    
    def handle_chat_bot(self):
        try:
            window = handle_chat_bot(self.window, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening chat bot window: {str(e)}")
            QMessageBox.critical(self.window, "Error", f"Failed to open LLM Assistance: {str(e)}")

    def handle_mitre_flow(self):
        try:
            if self.mitre_flow_window:
                self.mitre_flow_window.show()
                self.mitre_flow_window.raise_()
                self.mitre_flow_window.activateWindow()
                self.logger.info("MITRE Flow window shown from cache")
            else:
                window = open_mitre_flow_window(self.window)
                if window:
                    self.mitre_flow_window = window
                    self.track_child_window(window)
                    self.logger.info("MITRE Flow window created and cached")
        except Exception as e:
            self.logger.error(f"Error opening MITRE Attack Flow window: {str(e)}")
            QMessageBox.critical(self.window, "Error", f"Failed to open MITRE Attack Flow: {str(e)}")

    def open_custom_window(self):
        custom_window = QWidget(self.window) 
        custom_window.setWindowTitle("New Case Details")
        custom_window.setMinimumSize(600, 800)
        layout = QVBoxLayout()
        text_box = QTextEdit()
        layout.addWidget(text_box)
        submit_button = QPushButton("Submit")
        layout.addWidget(submit_button)

        def submit_data():
            data = text_box.toPlainText()
            custom_window.close()

        submit_button.clicked.connect(submit_data)
        custom_window.setLayout(layout)
        self.track_child_window(custom_window)
        custom_window.show()
    
    def open_api_settings(self):
        open_api_settings(self.window, self.db_path, self.logger, self.child_windows)

    def open_new_case_window(self):
        file_path, _ = QFileDialog.getSaveFileName(self.window, "Save New Case File", "New_Case.xlsx", "Excel files (*.xlsx)")
        if not file_path:
            QMessageBox.warning(self.window, "Warning", "No file selected. Case creation canceled.")
            return
        lock_acquired = False
        try:
            lock_status = self.acquire_file_lock(file_path)
            if lock_status is None:  
                return
            lock_acquired = True
            template_file = "sod.xlsx"
            if not os.path.exists(template_file):
                QMessageBox.critical(self.window, "Error", f"Template file '{template_file}' not found.")
                return
            shutil.copy(template_file, file_path)
            QMessageBox.information(self.window, "Success", f"New case created successfully and saved to {file_path}!")
            workbook = openpyxl.load_workbook(file_path)
            self.current_file_path = file_path
            self.current_workbook = workbook
            self.read_only_mode = False  
            file_status_label = self.window.ui.labelFileStatus
            if file_status_label:
                file_name = os.path.basename(file_path)
                file_status_label.setText(f"Loaded: {file_name}")
            else:
                self.logger.warning("labelFileStatus not found!")
            if self.tree_view:
                self.load_data_into_treeview(file_path=file_path, workbook=workbook)
            else:
                self.logger.warning("treeViewMain not found!")
        except Exception as e:
            self.logger.error(f"Error creating new case: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to create new case: {e}")
        finally:
            if lock_acquired and (not hasattr(self, 'current_workbook') or not self.current_workbook):
                self.release_file_lock()

    def load_data_into_treeview(self, file_path=None, workbook=None):
        lock_acquired = False
        progress = None
        try:
            if file_path is None or workbook is None:
                file_path, _ = QFileDialog.getOpenFileName(
                    self.window, "Open Excel File", "", "Excel Files (*.xlsx *.xls)"
                )
                if not file_path:
                    self.logger.info("No file selected")
                    return
                progress = QProgressBar()
                progress.setWindowTitle("Loading Excel File")
                progress.setGeometry(QRect(300, 300, 400, 30))
                progress.setWindowFlags(Qt.Window | Qt.WindowStaysOnTopHint)
                progress.setRange(0, 0)  
                progress.show()
                progress.setFormat("Loading file...")
                QApplication.processEvents()
                try:
                    lock_status = self.acquire_file_lock(file_path)
                    if lock_status is None:  
                        if progress:
                            progress.close()
                        return
                    lock_acquired = True
                    progress.setFormat("Opening workbook...")
                    QApplication.processEvents()
                    workbook = openpyxl.load_workbook(file_path, read_only=self.read_only_mode)
                except Exception as e:
                    if progress:
                        progress.close()
                    if lock_acquired:
                        self.release_file_lock()
                        lock_acquired = False
                    self.logger.error(f"Error loading Excel file: {e}")
                    QMessageBox.critical(self.window, "Error", f"Failed to load Excel file: {e}")
                    return

            self.current_workbook = workbook
            self.current_file_path = file_path
            self.window.current_workbook = workbook  
            self.window.current_file_path = file_path  
            if progress:
                progress.setFormat("Updating file status...")
                QApplication.processEvents()
            file_status_label = self.window.ui.labelFileStatus
            if file_status_label:
                file_name = os.path.basename(file_path)
                read_only_text = " [READ-ONLY]" if self.read_only_mode else ""
                file_status_label.setText(f"Loaded: {file_name}{read_only_text}")
            else:
                self.logger.warning("labelFileStatus not found!")
            sheet_dropdown = self.window.ui.comboBoxSheet
            if not sheet_dropdown:
                if progress:
                    progress.close()
                self.logger.warning("Sheet dropdown (comboBoxSheet) not found!")
                return
            if progress:
                progress.setFormat("Loading sheet names...")
                QApplication.processEvents()
            sheet_dropdown.clear()
            sheet_dropdown.blockSignals(True)
            sheet_dropdown.addItems(workbook.sheetnames)
            sheet_dropdown.blockSignals(False)
        
            def load_selected_sheet():
                selected_sheet_name = sheet_dropdown.currentText()
                if not selected_sheet_name:
                    self.logger.info("No sheet selected")
                    return
                
                show_progress = sys.platform != "darwin"
                sheet_progress = None
                
                if show_progress:
                    sheet_progress = QProgressBar()
                    sheet_progress.setWindowTitle(f"Loading Sheet: {selected_sheet_name}")
                    sheet_progress.setGeometry(QRect(300, 300, 400, 30))
                    sheet_progress.setWindowFlags(Qt.Window | Qt.WindowStaysOnTopHint)
                    sheet_progress.show()
                
                try:
                    current_workbook = self.current_workbook
                    if not current_workbook:
                        if sheet_progress:
                            sheet_progress.close()
                        return
                    self.current_sheet_name = selected_sheet_name
                    self.window.current_sheet_name = selected_sheet_name  
                    sheet = current_workbook[selected_sheet_name]
                    max_row = sheet.max_row
                    
                    if sheet_progress:
                        sheet_progress.setRange(0, max_row + 2)  
                        sheet_progress.setValue(0)
                        sheet_progress.setFormat("Reading headers...")
                        QApplication.processEvents()
                    
                    model = QStandardItemModel()
                    headers = [cell.value for cell in sheet[1]]  
                    model.setHorizontalHeaderLabels(headers)
                    
                    if sheet_progress:
                        sheet_progress.setValue(1)
                        sheet_progress.setFormat("Loading data...")
                        QApplication.processEvents()
                    
                    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):  
                        items = [QStandardItem(str(cell.value) if cell.value is not None else "") for cell in row]
                        model.appendRow(items)
                        if sheet_progress and (row_index % 100 == 0 or row_index == max_row):
                            sheet_progress.setValue(row_index)
                            sheet_progress.setFormat(f"Loading row {row_index}/{max_row}...")
                            QApplication.processEvents()
                    
                    if sheet_progress:
                        sheet_progress.setFormat("Finalizing...")
                        QApplication.processEvents()
                    
                    self.tree_view.setModel(model)
                    self.tree_view.setEditTriggers( 
                        QTreeView.NoEditTriggers if self.read_only_mode else 
                        (QTreeView.DoubleClicked | QTreeView.EditKeyPressed | QTreeView.AnyKeyPressed)
                    )
                    
                    for column in range(model.columnCount()):
                        self.tree_view.resizeColumnToContents(column)
                    
                    if sheet_progress:
                        sheet_progress.setValue(max_row + 2)
                        sheet_progress.close()
                except Exception as e:
                    if sheet_progress:
                        sheet_progress.close()
                    self.logger.error(f"Error loading sheet '{selected_sheet_name}': {e}")
                    self.tree_view.setModel(QStandardItemModel())
            sheet_dropdown.currentIndexChanged.connect(load_selected_sheet)
            if progress:
                progress.setFormat("Initializing sheet view...")
                QApplication.processEvents()
            sheet_dropdown.setCurrentIndex(0)
            load_selected_sheet()
            if progress:
                progress.close()
        except Exception as e:
            if progress:
                progress.close()
            self.logger.error(f"Error in load_data_into_treeview: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to load data into tree view: {e}")
            if lock_acquired and not workbook:  
                self.release_file_lock()
                
    def get_evidence_types_from_db(self):
        evidence_types = []
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT evidencetype FROM EvidenceType ORDER BY evidencetype")
            evidence_types = [row[0] for row in cursor.fetchall()]
            conn.close()
            if not evidence_types:
                evidence_types = []
        except sqlite3.Error as e:
            self.logger.error(f"Error fetching evidence types from database: {e}")
            evidence_types = []
        return evidence_types

    def edit_row(self, index):
        if self.read_only_mode:
            QMessageBox.information(self.window, "Read-Only Mode", "This file is open in read-only mode. You cannot edit its contents.")
            return
        if not index.isValid():
            QMessageBox.warning(self.window, "Warning", "No row selected. Please select a row to edit.")
            return
        visual_row_index = index.row()
        model = self.tree_view.model()
        if not model:
            QMessageBox.critical(self.window, "Error", "No model found for the tree view.")
            return
        row_data = []
        for col in range(model.columnCount()):
            cell_data = model.index(visual_row_index, col).data()
            row_data.append(cell_data)
        actual_row_index = None
        if self.current_workbook and self.current_sheet_name:
            sheet = self.current_workbook[self.current_sheet_name]
            for excel_row in range(2, sheet.max_row + 1):  
                excel_row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=excel_row, column=col).value
                    excel_row_data.append(str(cell_value) if cell_value is not None else "")
                if excel_row_data == [str(d) if d is not None else "" for d in row_data]:
                    actual_row_index = excel_row
                    break
        if actual_row_index is None:
            QMessageBox.critical(self.window, "Error", "Could not find matching data row.")
            return
        row_index = visual_row_index
        mitre_tactic_options = []
        mitre_techniques_by_tactic = {}  
        current_tactic = None
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT PID FROM mitre_techniques")
            mitre_tactic_options = [row[0] for row in cursor.fetchall()]
            for tactic in mitre_tactic_options:
                cursor.execute("SELECT ID FROM mitre_techniques WHERE PID = ?", (tactic,))
                mitre_techniques_by_tactic[tactic] = [row[0] for row in cursor.fetchall()]
            conn.close()
        except sqlite3.Error as e:
            QMessageBox.critical(self.window, "Error", f"Failed to fetch MITRE values: {e}")
        evidence_types = self.get_evidence_types_from_db()
        editor_widget = QWidget(self.window)
        editor_widget.setWindowTitle("Edit Row - Widget Editor")
        import sys
        if sys.platform == "darwin":  
            editor_widget.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint | 
                                      Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)
            editor_widget.setFixedSize(800, 600)
        else:  
            editor_widget.setWindowFlags(Qt.Window | Qt.WindowTitleHint | 
                                      Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint |
                                      Qt.WindowMaximizeButtonHint)
            editor_widget.setMinimumSize(500, 400)  
        grid_layout = QGridLayout()
        grid_layout.setContentsMargins(15, 15, 15, 15)
        grid_layout.setHorizontalSpacing(20)
        grid_layout.setVerticalSpacing(10)
        input_fields = []
        mitre_tactic_combo = None
        mitre_technique_combo = None
        existing_tactic_value = None
        existing_technique_value = None
        for col in range(model.columnCount()):
            header_text = model.headerData(col, Qt.Horizontal)
            cell_data = model.index(row_index, col).data()
            header_label = QLabel(f"{header_text}:")
            header_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if header_text and header_text.strip().lower() == "visualize":
                combo_box = QComboBox()
                combo_box.addItems(["Yes", "No"])
                if cell_data in ["Yes", "No"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "evidencetype":
                combo_box = QComboBox()
                combo_box.addItems(evidence_types)
                if cell_data in evidence_types:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "indicatortype":
                combo_box = QComboBox()
                combo_box.addItems(["IPAddress","UserName","FileName","FilePath","UserAgent","DomainName","JA3-JA3S","URL","Mutex","Other-Strings","EmailAddress","RegistryPath","GPO"])
                if cell_data in ["IPAddress","UserName","FileName","FilePath","UserAgent","DomainName","JA3-JA3S","URL","Mutex","Other-Strings","EmailAddress","RegistryPath","GPO"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "location":
                combo_box = QComboBox()
                combo_box.addItems(["On-Prem", "Unknown", "Cloud-Generic", "Cloud-Azure", "Cloud-AWS", "Clous-GCP"])
                if cell_data in ["On-Prem", "Unknown", "Cloud-Generic", "Cloud-Azure", "Cloud-AWS", "Clous-GCP"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "currentstatus":
                combo_box = QComboBox()
                combo_box.addItems(["Completed" , "In Progress", "On Hold", "Not Started"  ])
                if cell_data in ["Completed" , "In Progress", "On Hold", "Not Started"  ]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "priority":
                combo_box = QComboBox()
                combo_box.addItems(["High" ,"Medium" ,"Low" ])
                if cell_data in ["High" ,"Medium" ,"Low" ]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box                
            elif header_text and header_text.strip().lower() == "evidencecollected":
                combo_box = QComboBox()
                combo_box.addItems(["Yes", "No" ])
                if cell_data in ["Yes", "No" ]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box                
            elif header_text and header_text.strip().lower() == "targettype":
                combo_box = QComboBox()
                combo_box.addItems(["Machine", "Identity", "Others"])
                if cell_data in ["Machine", "Identity", "Others"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box               
            elif header_text and header_text.strip().lower() == "systemtype":
                combo_box = QComboBox()
                system_type_options = self.system_type_manager.get_system_type_options()
                combo_box.addItems([option[0] for option in system_type_options])
                if cell_data in [option[0] for option in system_type_options]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "accounttype":
                combo_box = QComboBox()
                combo_box.addItems(["Normal User Account - Local","Normal User Account - On-Prem AD","Normal User Account - Azure","Service Account", "Domain Admin", "Global Admin - Azure", "Service Principle - Azure", "Computer Account", "Local Administrator"])
                if cell_data in ["Normal User Account - Local","Normal User Account - On-Prem AD","Normal User Account - Azure","Service Account", "Domain Admin", "Global Admin - Azure", "Service Principle - Azure", "Computer Account", "Local Administrator"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text.strip().lower() == "entrypoint":
                combo_box = QComboBox()
                combo_box.addItems(["Yes", "No" ])
                if cell_data in ["Yes", "No" ]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box               
            elif header_text and (header_text.strip().lower() == "notes" or header_text.strip().lower() == "activity"):
                text_edit = QTextEdit()
                text_edit.setPlainText(cell_data if cell_data else "")
                input_field = text_edit
            elif header_text and header_text.strip().lower() in ["date added", "date updated", "date completed", "date requested", "date received"]:
                date_edit = QDateEdit()
                date_edit.setCalendarPopup(True)  
                if cell_data:
                    try:
                        date_edit.setDate(QDate.fromString(cell_data, "yyyy-MM-dd"))  
                    except Exception as e:
                        self.logger.error(f"Error parsing date: {e}")
                input_field = date_edit
            elif header_text and header_text == "TLP":
                combo_box = QComboBox()
                combo_box.addItems(["TLP-Red", "TLP-Amber", "TLP-Green", "TLP-Clear"])
                if cell_data in ["TLP-Red", "TLP-Amber", "TLP-Green", "TLP-Clear"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            elif header_text and header_text == "MITRE Tactic":
                existing_tactic_value = cell_data if cell_data else ""
                combo_box = QComboBox()
                combo_box.addItem("")  
                combo_box.addItems(mitre_tactic_options)
                if existing_tactic_value and existing_tactic_value in mitre_tactic_options:
                    combo_box.setCurrentText(existing_tactic_value)
                mitre_tactic_combo = combo_box
                input_field = combo_box
            elif header_text and header_text == "MITRE Techniques":
                existing_technique_value = cell_data if cell_data else ""
                combo_box = QComboBox()
                combo_box.addItem("")
                mitre_technique_combo = combo_box
                input_field = combo_box
            elif header_text and header_text == "<->":
                combo_box = QComboBox()
                combo_box.addItems([" ","->", "<-", "<->"])
                if cell_data in [" ","->", "<-", "<->"]:
                    combo_box.setCurrentText(cell_data)
                input_field = combo_box
            else:
                line_edit = QLineEdit()
                line_edit.setText(cell_data if cell_data else "")
                line_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                input_field = line_edit
            grid_layout.addWidget(header_label, col, 0)
            grid_layout.addWidget(input_field, col, 1)
            input_fields.append(input_field)
        if mitre_tactic_combo and mitre_technique_combo:
            current_tactic = mitre_tactic_combo.currentText()
            if current_tactic and current_tactic in mitre_techniques_by_tactic:
                available_techniques = mitre_techniques_by_tactic[current_tactic]
                mitre_technique_combo.addItems(available_techniques)
                if existing_technique_value:
                    if existing_technique_value in available_techniques:
                        mitre_technique_combo.setCurrentText(existing_technique_value)
                    else:
                        mitre_technique_combo.addItem(existing_technique_value)
                        mitre_technique_combo.setCurrentText(existing_technique_value)

            def update_techniques(index):
                selected_tactic = mitre_tactic_combo.currentText()
                current_technique = mitre_technique_combo.currentText()
                mitre_technique_combo.clear()
                mitre_technique_combo.addItem("")
                if selected_tactic and selected_tactic in mitre_techniques_by_tactic:
                    techniques = mitre_techniques_by_tactic[selected_tactic]
                    mitre_technique_combo.addItems(techniques)
                    if current_technique and current_technique in techniques:
                        mitre_technique_combo.setCurrentText(current_technique)
            mitre_tactic_combo.currentIndexChanged.connect(update_techniques)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(20)
        save_button = QPushButton("Save")
        cancel_button = QPushButton("Cancel")
        button_layout.addStretch()
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        main_layout = QVBoxLayout(editor_widget)
        main_layout.addLayout(grid_layout)
        main_layout.addLayout(button_layout)
        editor_widget.setLayout(main_layout)

        def save_changes():
            try:
                if not self.current_workbook or not self.current_sheet_name:
                    QMessageBox.critical(self.window, "Error", "No workbook or sheet loaded to save changes.")
                    return
                workbook = self.current_workbook
                sheet = workbook[self.current_sheet_name]
                for col, field in enumerate(input_fields):
                    if isinstance(field, QComboBox):
                        new_text = field.currentText()
                    elif isinstance(field, QTextEdit):
                        new_text = field.toPlainText()
                    elif isinstance(field, QDateEdit):
                        new_text = field.date().toString("yyyy-MM-dd")  
                    else:
                        new_text = field.text()
                    sheet.cell(row=actual_row_index, column=col + 1, value=new_text)
                    model.setData(model.index(row_index, col), new_text)
                workbook.save(self.current_file_path)
                editor_widget.close()
            except Exception as e:
                self.logger.error(f"Error in edit_row save_changes: {e}")
                QMessageBox.critical(self.window, "Error", f"Failed to save changes: {e}")
        save_button.clicked.connect(save_changes)
        cancel_button.clicked.connect(editor_widget.close)
        editor_widget.show()
        self.track_child_window(editor_widget)

    def display_bookmarks_kb(self):
        window = display_bookmarks_kb(self, self.db_path)
        self.track_child_window(window)

    def display_event_id_kb(self):
        try:
            window = display_event_id_kb(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error in display_event_id_kb: {e}")
            traceback.print_exc()
            QMessageBox.critical(self.window, "Error", f"Failed to open Event ID Knowledge Base: {str(e)}")

    def display_lolbas_kb(self):
        try:
            window = display_lolbas_kb(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error in display_lolbas_kb: {e}")
            traceback.print_exc()
            QMessageBox.critical(self.window, "Error", f"Failed to open LOLBAS Knowledge Base: {str(e)}")

    def entra_appid(self):
        window = open_entra_lookup_window(self.window, self.db_path)
        self.track_child_window(window)

    def handle_cve_lookup(self):
        try:
            window = open_cve_window(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening CVE lookup window: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open CVE lookup window: {e}")

    def handle_stix_export(self):
        if not self.check_excel_loaded():
            return
            
        try:
            if 'Indicators' not in self.current_workbook.sheetnames:
                QMessageBox.warning(
                    self.window, 
                    "Warning", 
                    "No 'Indicators' sheet found in the loaded Excel file.\n\n"
                    "Please ensure your Excel file contains an 'Indicators' sheet with columns like:\n"
                    "- name\n"
                    "- description\n" 
                    "- pattern\n"
                    "- indicator_types\n"
                    "- confidence (optional)\n"
                    "- valid_until (optional)"
                )
                return
            try:
                stix_bundle = convert_indicators_to_stix(
                    excel_file_path=self.current_file_path,
                    sheet_name="Indicators"
                )
                self.show_stix_json_window(stix_bundle)
                
            except Exception as e:
                QMessageBox.critical(
                    self.window,
                    "STIX Export Error",
                    f"Error during STIX conversion:\n{str(e)}\n\n"
                    f"Please check that your 'Indicator' sheet has the correct format."
                )
                self.logger.error(f"STIX export error: {e}")
                
        except Exception as e:
            QMessageBox.critical(
                self.window,
                "STIX Export Error",
                f"Unexpected error during STIX export:\n{str(e)}"
            )
            self.logger.error(f"Unexpected STIX export error: {e}")

    def show_stix_json_window(self, stix_bundle):
        import json
        json_window = QWidget(self.window)
        json_window.setWindowTitle("STIX Export - JSON Data")
        import sys
        if sys.platform == "darwin":  
            json_window.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint | 
                                    Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)
            json_window.setFixedSize(1000, 700)
        else:  
            json_window.setWindowFlags(Qt.Window | Qt.WindowTitleHint | 
                                    Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint |
                                    Qt.WindowMaximizeButtonHint)
            json_window.setMinimumSize(800, 600)
        main_layout = QVBoxLayout(json_window)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)
        info_label = QLabel(f"Generated {len(stix_bundle.get('objects', []))} STIX objects")
        info_label.setFont(QFont("Arial", 10))
        info_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(info_label)
        json_text = QTextEdit()
        json_text.setFont(QFont("Courier", 9))
        json_text.setReadOnly(True)
        try:
            formatted_json = json.dumps(stix_bundle, indent=2, ensure_ascii=False)
            json_text.setPlainText(formatted_json)
        except Exception as e:
            json_text.setPlainText(f"Error formatting JSON: {str(e)}")
        scroll_area = QScrollArea()
        scroll_area.setWidget(json_text)
        scroll_area.setWidgetResizable(True)
        main_layout.addWidget(scroll_area)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        copy_button = QPushButton("Copy to Clipboard")
        copy_button.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")
        save_button = QPushButton("Save to File")
        save_button.setStyleSheet("background-color: #2196F3; color: white; padding: 8px;")
        close_button = QPushButton("Close")
        close_button.setStyleSheet("background-color: #f44336; color: white; padding: 8px;")
        button_layout.addStretch()
        button_layout.addWidget(copy_button)
        button_layout.addWidget(save_button)
        button_layout.addWidget(close_button)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        
        def copy_to_clipboard():
            try:
                clipboard = QApplication.clipboard()
                clipboard.setText(json_text.toPlainText())
                QMessageBox.information(json_window, "Success", "JSON data copied to clipboard!")
            except Exception as e:
                QMessageBox.critical(json_window, "Error", f"Failed to copy to clipboard: {str(e)}")
        
        def save_to_file():
            try:
                from datetime import datetime
                base_name = os.path.splitext(os.path.basename(self.current_file_path))[0]
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                default_filename = f"{base_name}_STIX_{timestamp}.json"
                
                file_path, _ = QFileDialog.getSaveFileName(
                    json_window,
                    "Save STIX JSON File",
                    default_filename,
                    "JSON Files (*.json);;All Files (*)"
                )
                
                if file_path:
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(json_text.toPlainText())
                    QMessageBox.information(json_window, "Success", f"STIX data saved to:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(json_window, "Error", f"Failed to save file: {str(e)}")
        
        copy_button.clicked.connect(copy_to_clipboard)
        save_button.clicked.connect(save_to_file)
        close_button.clicked.connect(json_window.close)
        json_window.show()
        self.track_child_window(json_window)

    def handle_ransomware_kb(self):
        try:
            window = open_ransomware_kb_window(self)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening Ransomware Knowledge Base window: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open Ransomware Knowledge Base window: {e}")

    def handle_defend_mapping(self):
        if self.check_excel_loaded():
            try:
                if 'Timeline' in self.current_workbook.sheetnames:
                    window = open_defend_window(self.window, self.current_file_path)
                    self.track_child_window(window)
                else:
                    QMessageBox.warning(self.window, "Missing Sheet", "The required 'Timeline' sheet was not found in this workbook.")
            except Exception as e:
                self.logger.error(f"Error in D3FEND mapping: {e}")
                QMessageBox.critical(self.window, "D3FEND Mapping Error", f"An error occurred while opening the D3FEND mapping window:\n\n{str(e)}")

    def handle_ip_lookup(self):
        try:
            window = open_ip_lookup_window(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening IP lookup window: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open IP lookup window: {e}")
            
    def handle_domain_lookup(self):
        try:
            window = open_domain_lookup_window(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening Domain lookup window: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open Domain lookup window: {e}")

    def handle_hash_lookup(self):
        try:
            window = open_hash_lookup_window(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening Hash lookup window: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open Hash lookup window: {e}")

    def handle_email_lookup(self):
        try:
            window = open_email_lookup_window(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening Email lookup window: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open Email lookup window: {e}")

    def handle_markdown_editor(self):
        try:
            window = handle_markdown_editor(self.window)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening Markdown Editor interface: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open Markdown Editor interface: {e}")

    def add_new_row(self):
        if self.read_only_mode:
            QMessageBox.information(self.window, "Read-Only Mode", "This file is open in read-only mode. You cannot add new rows.")
            return
        if not self.current_sheet_name or not self.current_workbook:
            QMessageBox.warning(self.window, "Warning", "Please select a sheet first.")
            return
        mitre_tactic_options = []
        mitre_techniques_by_tactic = {}  
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT PID FROM mitre_techniques")
            mitre_tactic_options = [row[0] for row in cursor.fetchall()]
            for tactic in mitre_tactic_options:
                cursor.execute("SELECT ID FROM mitre_techniques WHERE PID = ?", (tactic,))
                mitre_techniques_by_tactic[tactic] = [row[0] for row in cursor.fetchall()]
            conn.close()
        except sqlite3.Error as e:
            self.logger.error(f"Error fetching MITRE values: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to fetch MITRE values: {e}")
            return
        evidence_types = self.get_evidence_types_from_db()
        sheet = self.current_workbook[self.current_sheet_name]
        headers = [cell.value for cell in sheet[1]]  
        add_row_window = QWidget(self.window)
        add_row_window.setWindowTitle("Add New Row - Widget Editor")
        import sys
        if sys.platform == "darwin":  
            add_row_window.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint | 
                                      Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)
            add_row_window.setFixedSize(800, 600)
        else:  
            add_row_window.setWindowFlags(Qt.Window | Qt.WindowTitleHint | 
                                      Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint |
                                      Qt.WindowMaximizeButtonHint)
            add_row_window.setMinimumSize(500, 400)  
        main_layout = QVBoxLayout(add_row_window)
        grid_layout = QGridLayout()
        grid_layout.setContentsMargins(15, 15, 15, 15)
        grid_layout.setHorizontalSpacing(20)
        grid_layout.setVerticalSpacing(10)
        input_fields = []
        mitre_tactic_combo = None
        mitre_technique_combo = None
        current_tactic = None
        for col, header in enumerate(headers):
            header_label = QLabel(f"{header}:")
            header_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            grid_layout.addWidget(header_label, col, 0)
            if header and header.strip().lower() == "visualize":
                combo_box = QComboBox()
                combo_box.addItems(["Yes", "No"])
                input_field = combo_box
            elif header and header.strip().lower() == "evidencetype":
                combo_box = QComboBox()
                combo_box.addItems(evidence_types)
                input_field = combo_box
            elif header and header.strip().lower() == "indicatortype":
                combo_box = QComboBox()
                combo_box.addItems(["IPAddress","UserName","FileName","FilePath","UserAgent","DomainName","JA3-JA3S","URL","Mutex","Other-Strings","EmailAddress","RegistryPath","GPO"])
                input_field = combo_box
            elif header and header.strip().lower() == "accounttype":
                combo_box = QComboBox()
                combo_box.addItems(["Normal User Account - Local","Normal User Account - On-Prem AD","Normal User Account - Azure","Service Account", "Domain Admin", "Global Admin - Azure", "Service Principle - Azure", "Computer Account", "Local Administrator"])
                input_field = combo_box
            elif header and header.strip().lower() == "systemtype":
                combo_box = QComboBox()
                system_type_options = self.system_type_manager.get_system_type_options()
                combo_box.addItems([option[0] for option in system_type_options])
                input_field = combo_box
            elif header and header.strip().lower() == "location":
                combo_box = QComboBox()
                combo_box.addItems(["On-Prem", "Unknown", "Cloud-Generic", "Cloud-Azure", "Cloud-AWS", "Clous-GCP"])
                input_field = combo_box
            elif header and header.strip().lower() == "currentstatus":
                combo_box = QComboBox()
                combo_box.addItems(["Completed" , "In Progress", "On Hold", "Not Started"  ])
                input_field = combo_box                
            elif header and header.strip().lower() == "priority":
                combo_box = QComboBox()
                combo_box.addItems(["High" ,"Medium" ,"Low" ])
                input_field = combo_box                
            elif header and header.strip().lower() == "evidencecollected":
                combo_box = QComboBox()
                combo_box.addItems(["Yes", "No" ])
                input_field = combo_box                
            elif header and header.strip().lower() == "entrypoint":
                combo_box = QComboBox()
                combo_box.addItems(["Yes", "No" ])
                input_field = combo_box                
            elif header and header.strip().lower() == "targettype":
                combo_box = QComboBox()
                combo_box.addItems(["Machine", "Identity", "Others"])
                input_field = combo_box                
            elif header and (header.strip().lower() == "notes" or header.strip().lower() == "activity"):
                text_edit = QTextEdit()
                input_field = text_edit
            elif header and "timestamp" in header.strip().lower() and "utc" in header.strip().lower():
                line_edit = QLineEdit()
                line_edit.setPlaceholderText("YYYY-MM-DD HH:MM:SS")
                line_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                input_field = line_edit
            elif header and header.strip().lower() in ["date added", "date updated", "date completed", "date requested", "date received"]:
                date_edit = QDateEdit()
                date_edit.setCalendarPopup(True)  
                date_edit.setDate(QDate.currentDate())  
                input_field = date_edit
            elif header == "MITRE Tactic":
                combo_box = QComboBox()
                combo_box.addItem("")  
                combo_box.addItems(mitre_tactic_options)
                mitre_tactic_combo = combo_box
                input_field = combo_box
            elif header == "MITRE Techniques":
                combo_box = QComboBox()
                combo_box.addItem("")  
                mitre_technique_combo = combo_box
                input_field = combo_box
            elif header == "TLP":
                combo_box = QComboBox()
                combo_box.addItems(["TLP-Red", "TLP-Amber", "TLP-Green", "TLP-Clear"])
                input_field = combo_box
            elif header == "<->":
                combo_box = QComboBox()
                combo_box.addItems([" ","->", "<-", "<->"])
                input_field = combo_box
            else:
                line_edit = QLineEdit()
                line_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                input_field = line_edit
            grid_layout.addWidget(header_label, col, 0)
            grid_layout.addWidget(input_field, col, 1)
            input_fields.append(input_field)
        if mitre_tactic_combo and mitre_technique_combo:
            current_technique = mitre_technique_combo.currentText()
            
            def update_techniques(tactic_index):
                current_selection = mitre_technique_combo.currentText()
                selected_tactic = mitre_tactic_combo.currentText()
                mitre_technique_combo.clear()
                mitre_technique_combo.addItem("")  
                if selected_tactic:
                    techniques = mitre_techniques_by_tactic.get(selected_tactic, [])
                    mitre_technique_combo.addItems(techniques)
                    if current_selection and current_selection in techniques:
                        mitre_technique_combo.setCurrentText(current_selection)
                else:
                    if current_selection:
                        mitre_technique_combo.addItem(current_selection)
                        mitre_technique_combo.setCurrentText(current_selection)
            mitre_tactic_combo.currentIndexChanged.connect(update_techniques)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(20)
        save_button = QPushButton("Save")
        cancel_button = QPushButton("Cancel")
        button_layout.addStretch()
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        main_layout.addLayout(grid_layout)
        main_layout.addLayout(button_layout)
    
        def save_new_row():
            try:
                if not self.current_workbook or not self.current_sheet_name:
                    QMessageBox.critical(self.window, "Error", "No workbook or sheet loaded to save changes.")
                    return
                sheet = self.current_workbook[self.current_sheet_name]
                new_row_data = []
                for field in input_fields:
                    if isinstance(field, QComboBox):
                        new_text = field.currentText()
                    elif isinstance(field, QTextEdit):
                        new_text = field.toPlainText()
                    elif isinstance(field, QDateEdit):
                        new_text = field.date().toString("yyyy-MM-dd")  
                    else:
                        new_text = field.text()
                    new_row_data.append(new_text)
                sheet.append(new_row_data)
                self.current_workbook.save(self.current_file_path)
                self.load_sheet()
                add_row_window.close()
            except Exception as e:
                self.logger.error(f"Failed to save changes: {e}")
                QMessageBox.critical(add_row_window, "Error", f"Failed to save changes: {e}")
        save_button.clicked.connect(save_new_row)
        cancel_button.clicked.connect(add_row_window.close)
        add_row_window.show()
        self.track_child_window(add_row_window)

    def delete_row(self):
        if self.read_only_mode:
            QMessageBox.information(self.window, "Read-Only Mode", "This file is open in read-only mode. You cannot delete rows.")
            return
        if not self.current_sheet_name or not self.current_workbook:
            QMessageBox.warning(self.window, "Warning", "Please select a sheet.")
            return
        if not self.tree_view or not self.tree_view.model():
            QMessageBox.warning(self.window, "Warning", "Tree view not available.")
            return
        selected_indices = self.tree_view.selectionModel().selectedRows()
        if not selected_indices:
            QMessageBox.warning(self.window, "Warning", "Please select rows to delete.")
            return
        model = self.tree_view.model()
        actual_row_indices = []
        for index in selected_indices:
            row_data = []
            for col in range(model.columnCount()):
                cell_data = model.index(index.row(), col).data()
                row_data.append(cell_data)
            actual_row = None
            sheet = self.current_workbook[self.current_sheet_name]
            for excel_row in range(2, sheet.max_row + 1):  
                excel_row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=excel_row, column=col).value
                    excel_row_data.append(str(cell_value) if cell_value is not None else "")
                if excel_row_data == [str(d) if d is not None else "" for d in row_data]:
                    actual_row = excel_row
                    break
            
            if actual_row is not None:
                actual_row_indices.append(actual_row)
        
        if not actual_row_indices:
            QMessageBox.warning(self.window, "Warning", "Could not find matching data rows.")
            return
        actual_row_indices.sort(reverse=True)
        
        confirm = QMessageBox.question(self.window, "Confirm Deletion", f"Are you sure you want to delete {len(actual_row_indices)} row(s)?", QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.No:
            return
        sheet = self.current_workbook[self.current_sheet_name]
        try:
            for row_idx in actual_row_indices:
                sheet.delete_rows(row_idx, 1)  
            self.current_workbook.save(self.current_file_path)
            QMessageBox.information(self.window, "Success", "Rows deleted successfully!")
            self.load_sheet()
        except PermissionError:
            self.logger.error("Permission error: Cannot write to file")
            QMessageBox.critical(self.window, "Error", "Cannot save file. Ensure it is not open in another program.")
        except Exception as e:
            self.logger.error(f"Error deleting rows: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to delete rows: {e}")

    def load_sheet(self):
        if not self.current_workbook or not self.current_sheet_name or not self.tree_view:
            self.logger.error("Cannot load sheet: Missing workbook, sheet name, or tree view")
            return
        try:
            sheet = self.current_workbook[self.current_sheet_name]
            model = QStandardItemModel()
            headers = []
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    headers.append(str(header))
                else:
                    headers.append(f"Column {col}")
            model.setHorizontalHeaderLabels(headers)
            for row_idx in range(2, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                items = [QStandardItem(value) for value in row_data]
                model.appendRow(items)
            self.tree_view.setModel(model)
            self.tree_view.setEditTriggers(QTreeView.NoEditTriggers)
            for col_idx, header in enumerate(headers):
                self.tree_view.setColumnWidth(col_idx, 100)
        except Exception as e:
            self.logger.error(f"Error loading sheet: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to load sheet: {e}")

    def run(self):
        self.app.aboutToQuit.connect(self.application_cleanup)
        return self.app.exec()

    def list_systems(self):
        if not self.current_workbook:
            QMessageBox.warning(self.window, "Warning", "No workbook loaded. Please load a workbook first.")
            return
        if 'Timeline' not in self.current_workbook.sheetnames:
            QMessageBox.warning(self.window, "Missing Sheet", "The 'Timeline' sheet was not found in the current workbook.")
            return
        try:
            sheet = self.current_workbook['Timeline']
            event_system_col = None
            remote_system_col = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header == "Event System":
                    event_system_col = col
                elif header == "Remote System":
                    remote_system_col = col;
            if event_system_col is None and remote_system_col is None:
                QMessageBox.warning(self.window, "Missing Columns", "Neither 'Event System' nor 'Remote System' columns found in the Timeline sheet.")
                return
            systems = set()
            for row in range(2, sheet.max_row + 1): 
                if event_system_col:
                    event_system = sheet.cell(row=row, column=event_system_col).value
                    if event_system:
                        systems.add(str(event_system).strip())
                if remote_system_col:
                    remote_system = sheet.cell(row=row, column=remote_system_col).value
                    if remote_system:
                        systems.add(str(remote_system).strip())
            sorted_systems = sorted(list(systems))
            systems_window = QWidget(self.window)  
            systems_window.setWindowTitle("Unique Systems")
            systems_window.resize(400, 500)
            systems_window.setWindowFlags(Qt.Window | Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
            layout = QVBoxLayout(systems_window)
            label = QLabel(f"Found {len(sorted_systems)} unique systems:")
            label.setFont(QFont(self.get_sans_serif_font(), 10, QFont.Bold))
            layout.addWidget(label)
            list_widget = QTreeView()
            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["System Name"])
            for system in sorted_systems:
                item = QStandardItem(system)
                model.appendRow(item)
            list_widget.setModel(model)
            self.apply_standard_treeview_styling(list_widget)
            list_widget.setEditTriggers(QTreeView.NoEditTriggers)
            header = list_widget.header()
            header.setSectionResizeMode(0, QHeaderView.Stretch)
            list_widget.resizeColumnToContents(0)
            layout.addWidget(list_widget, 1) 
            button_layout = QHBoxLayout()
            close_button = QPushButton("Close")
            close_button.clicked.connect(systems_window.close)
            button_layout.addStretch()
            button_layout.addWidget(close_button)
            button_layout.addStretch()
            layout.addLayout(button_layout)
            self.child_windows.append(systems_window)
            systems_window.show()
        except Exception as e:
            self.logger.error(f"Error listing systems: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to list systems: {e}")
        
    def close_systems_window(self):
        if hasattr(self, 'systems_window'):
            self.systems_window.close()
            self.systems_window = None

    def list_users(self):
        if not self.current_workbook:
            QMessageBox.warning(self.window, "Warning", "No workbook loaded. Please load a workbook first.")
            return
        if 'Timeline' not in self.current_workbook.sheetnames:
            QMessageBox.warning(self.window, "Missing Sheet", "The 'Timeline' sheet was not found in the current workbook.")
            return
        try:
            sheet = self.current_workbook['Timeline']
            suspect_account_col = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header == "Suspect Account":
                    suspect_account_col = col
                    break
            if suspect_account_col is None:
                QMessageBox.warning(self.window, "Missing Column", "'Suspect Account' column not found in the Timeline sheet.")
                return
            users = set()
            for row in range(2, sheet.max_row + 1):  
                user = sheet.cell(row=row, column=suspect_account_col).value
                if user:
                    users.add(str(user).strip())
            sorted_users = sorted(list(users))
            users_window = QWidget(self.window)  
            users_window.setWindowTitle("Unique Users")
            users_window.resize(400, 500)
            users_window.setWindowFlags(Qt.Window | Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
            layout = QVBoxLayout(users_window)
            label = QLabel(f"Found {len(sorted_users)} unique users:")
            label.setFont(QFont(self.get_sans_serif_font(), 10, QFont.Bold))
            layout.addWidget(label)
            list_widget = QTreeView()
            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["Username"])
            for user in sorted_users:
                item = QStandardItem(user)
                model.appendRow(item)
            list_widget.setModel(model)
            self.apply_standard_treeview_styling(list_widget)
            list_widget.setEditTriggers(QTreeView.NoEditTriggers)
            header = list_widget.header()
            header.setSectionResizeMode(0, QHeaderView.Stretch)
            list_widget.resizeColumnToContents(0)
            layout.addWidget(list_widget, 1)  
            button_layout = QHBoxLayout()
            close_button = QPushButton("Close")
            close_button.clicked.connect(users_window.close)
            button_layout.addStretch()
            button_layout.addWidget(close_button)
            button_layout.addStretch()
            layout.addLayout(button_layout)
            self.child_windows.append(users_window)
            users_window.show()
        except Exception as e:
            traceback_str = traceback.format_exc()
            self.logger.error(f"Error listing users: {e}")
            self.logger.error(f"Traceback: {traceback_str}")
            QMessageBox.critical(self.window, "Error", f"Failed to list users: {e}")

    def close_users_window(self):
        if hasattr(self, 'users_window'):
            self.users_window.close()
            self.users_window = None

    def defang(self):
        if not self.check_excel_loaded():
            return
        try:
            defanged_file_path, _ = QFileDialog.getSaveFileName(self.window, "Save Defanged Excel File", os.path.splitext(self.current_file_path)[0] + "_defanged.xlsx", "Excel files (*.xlsx)")
            if not defanged_file_path:
                return
            lock_acquired = False
            try:
                lock_status = self.acquire_file_lock(defanged_file_path)
                if lock_status is None:  
                    return
                lock_acquired = True
                progress = QProgressBar()
                progress.setWindowTitle("Defanging File")
                progress.setGeometry(QRect(300, 300, 400, 30))
                progress.setWindowFlags(Qt.Window | Qt.WindowStaysOnTopHint)
                progress.setRange(0, len(self.current_workbook.sheetnames))
                progress.setValue(0)
                progress.show()
                
                def progress_callback(current, total, message):
                    progress.setValue(current)
                    progress.setWindowTitle(message)
                    QApplication.processEvents()
                
                defang_excel_file(self.current_file_path, defanged_file_path, progress_callback)
                
                progress.close()
                QMessageBox.information(self.window, "Defanging Complete", f"File has been defanged and saved to:\n{defanged_file_path}")
                reply = QMessageBox.question(self.window, "Load Defanged File", "Do you want to load the defanged file?", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                if reply == QMessageBox.Yes:
                    self.release_file_lock()
                    lock_acquired = False
                    self.load_data_into_treeview(file_path=defanged_file_path)
            finally:
                if lock_acquired:
                    self.release_file_lock()
        except Exception as e:
            self.logger.error(f"Error during defanging: {e}")
            traceback.print_exc()
            QMessageBox.critical(self.window, "Defanging Error", f"An error occurred while defanging the file:\n{str(e)}")

    def edit_row_from_context(self):
        selected_indices = self.tree_view.selectionModel().selectedRows()
        if selected_indices:
            self.edit_row(selected_indices[0])
    
    def duplicate_row(self):
        if self.read_only_mode:
            QMessageBox.information(self.window, "Read-Only Mode", "This file is open in read-only mode. You cannot duplicate rows.")
            return
        if not self.current_sheet_name or not self.current_workbook:
            QMessageBox.warning(self.window, "Warning", "Please select a sheet first.")
            return
        selected_indices = self.tree_view.selectionModel().selectedRows()
        if not selected_indices:
            QMessageBox.warning(self.window, "Warning", "Please select rows to duplicate.")
            return
        try:
            sheet = self.current_workbook[self.current_sheet_name]
            model = self.tree_view.model()
            if not model:
                QMessageBox.critical(self.window, "Error", "No model found for the tree view.")
                return
            actual_row_indices = []
            for index in selected_indices:
                row_data = []
                for col in range(model.columnCount()):
                    cell_data = model.index(index.row(), col).data()
                    row_data.append(cell_data)
                actual_row = None
                for excel_row in range(2, sheet.max_row + 1):  
                    excel_row_data = []
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=excel_row, column=col).value
                        excel_row_data.append(str(cell_value) if cell_value is not None else "")
                    if excel_row_data == [str(d) if d is not None else "" for d in row_data]:
                        actual_row = excel_row
                        break
                if actual_row is not None:
                    actual_row_indices.append(actual_row)
            if not actual_row_indices:
                QMessageBox.warning(self.window, "Warning", "Could not find matching data rows.")
                return
            actual_row_indices.sort(reverse=True)
            for row_idx in actual_row_indices:
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col).value
                    row_data.append(cell_value)
                sheet.insert_rows(row_idx + 1)
                for col, value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx + 1, column=col, value=value)
            self.current_workbook.save(self.current_file_path)
            QMessageBox.information(self.window, "Success", f"Duplicated {len(actual_row_indices)} row(s) successfully!")
            self.load_sheet()
        except Exception as e:
            self.logger.error(f"Error duplicating rows: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to duplicate rows: {str(e)}")
    
    def copy_row_data(self):
        selected_indices = self.tree_view.selectionModel().selectedRows()
        if not selected_indices:
            QMessageBox.warning(self.window, "Warning", "Please select rows to copy.")
            return
        try:
            model = self.tree_view.model()
            if not model:
                QMessageBox.critical(self.window, "Error", "No model found for the tree view.")
                return
            clipboard_data = []
            for index in selected_indices:
                row_data = []
                for col in range(model.columnCount()):
                    cell_index = model.index(index.row(), col)
                    cell_data = model.data(cell_index)
                    row_data.append(str(cell_data) if cell_data is not None else "")
                clipboard_data.append("\t".join(row_data))
            clipboard_text = "\n".join(clipboard_data)
            clipboard = QApplication.clipboard()
            clipboard.setText(clipboard_text)
            QMessageBox.information(self.window, "Success", f"Copied {len(selected_indices)} row(s) to clipboard!")
        except Exception as e:
            self.logger.error(f"Error copying row data: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to copy row data: {str(e)}")
    
    def export_selected_to_excel(self):
        selected_indices = self.tree_view.selectionModel().selectedRows()
        if not selected_indices:
            QMessageBox.warning(self.window, "Warning", "Please select rows to export.")
            return
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self.window, "Export Selected Rows", "", "Excel Files (*.xlsx)"
            )
            if not file_path:
                return
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Exported Data"
            if self.current_sheet_name and self.current_workbook:
                source_sheet = self.current_workbook[self.current_sheet_name]
                headers = [cell.value for cell in source_sheet[1]]
                ws.append(headers)
                for index in selected_indices:
                    row_data = []
                    for col in range(1, source_sheet.max_column + 1):
                        cell_value = source_sheet.cell(row=index.row() + 2, column=col).value
                        row_data.append(cell_value)
                    ws.append(row_data)
            wb.save(file_path)
            QMessageBox.information(self.window, "Success", f"Exported {len(selected_indices)} row(s) to {file_path}")
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to export to Excel: {str(e)}")
    
    def export_selected_to_csv(self):
        selected_indices = self.tree_view.selectionModel().selectedRows()
        if not selected_indices:
            QMessageBox.warning(self.window, "Warning", "Please select rows to export.")
            return
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self.window, "Export Selected Rows", "", "CSV Files (*.csv)"
            )
            if not file_path:
                return
            
            import csv
            if self.current_sheet_name and self.current_workbook:
                source_sheet = self.current_workbook[self.current_sheet_name]
                headers = [str(cell.value) if cell.value else "" for cell in source_sheet[1]]
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(headers)
                    for index in selected_indices:
                        row_data = []
                        for col in range(1, source_sheet.max_column + 1):
                            cell_value = source_sheet.cell(row=index.row() + 2, column=col).value
                            row_data.append(str(cell_value) if cell_value else "")
                        writer.writerow(row_data)
            QMessageBox.information(self.window, "Success", f"Exported {len(selected_indices)} row(s) to {file_path}")
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to export to CSV: {str(e)}")


    def handle_add_system_type(self):
        from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QMessageBox
        from PySide6.QtCore import Qt
        import os
        import shutil
        
        dialog = QWidget(self.window)
        dialog.setWindowTitle("Add System Type")
        dialog.setFixedSize(500, 300)
        dialog.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint | Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)
        
        layout = QVBoxLayout()
        
        title_label = QLabel("Add New System Type")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title_label)
        
        name_layout = QHBoxLayout()
        name_label = QLabel("System Type Name:")
        name_label.setMinimumWidth(150)
        name_input = QLineEdit()
        name_input.setPlaceholderText("e.g., Server-Cloud, Gateway-IDS")
        name_layout.addWidget(name_label)
        name_layout.addWidget(name_input)
        layout.addLayout(name_layout)
        
        icon_layout = QHBoxLayout()
        icon_label = QLabel("Icon File:")
        icon_label.setMinimumWidth(150)
        icon_input = QLineEdit()
        icon_input.setPlaceholderText("e.g., cloud_server.png")
        icon_browse_btn = QPushButton("Browse...")
        icon_layout.addWidget(icon_label)
        icon_layout.addWidget(icon_input)
        icon_layout.addWidget(icon_browse_btn)
        layout.addLayout(icon_layout)
        
        color_layout = QHBoxLayout()
        color_label = QLabel("Fallback Color:")
        color_label.setMinimumWidth(150)
        color_input = QLineEdit()
        color_input.setPlaceholderText("#00bfff")
        color_input.setText("#808080")
        color_layout.addWidget(color_label)
        color_layout.addWidget(color_input)
        layout.addLayout(color_layout)
        
        info_label = QLabel("Note: Icon files will be copied to the 'images/' directory")
        info_label.setStyleSheet("color: #666; font-size: 12px; margin-top: 10px;")
        layout.addWidget(info_label)
        
        button_layout = QHBoxLayout()
        add_button = QPushButton("Add System Type")
        add_button.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")
        cancel_button = QPushButton("Cancel")
        cancel_button.setStyleSheet("background-color: #f44336; color: white; padding: 8px;")
        button_layout.addWidget(add_button)
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        dialog.setLayout(layout)
        
        def browse_icon_file():
            file_path, _ = QFileDialog.getOpenFileName(
                dialog,
                "Select Icon File",
                "",
                "Image Files (*.png *.jpg *.jpeg *.svg *.bmp *.gif);;All Files (*)"
            )
            if file_path:
                filename = os.path.basename(file_path)
                icon_input.setText(filename)
                
                images_dir = os.path.join(os.path.dirname(__file__), "images")
                os.makedirs(images_dir, exist_ok=True)
                dest_path = os.path.join(images_dir, filename)
                
                try:
                    shutil.copy2(file_path, dest_path)
                    QMessageBox.information(dialog, "Success", f"Icon file copied to: {dest_path}")
                except Exception as e:
                    QMessageBox.warning(dialog, "Warning", f"Failed to copy icon file: {str(e)}")
        
        def add_system_type():
            name = name_input.text().strip()
            icon_filename = icon_input.text().strip()
            fallback_color = color_input.text().strip()
            
            if not name:
                QMessageBox.warning(dialog, "Warning", "Please enter a System Type Name.")
                return
            
            if not fallback_color:
                fallback_color = "#808080"
            
            if icon_filename:
                images_dir = os.path.join(os.path.dirname(__file__), "images")
                icon_path = os.path.join(images_dir, icon_filename)
                if not os.path.exists(icon_path):
                    QMessageBox.warning(dialog, "Warning", f"Icon file not found: {icon_path}")
                    return
            
            try:
                success = self.system_type_manager.add_system_type(
                    name=name,
                    display_name=name.replace('-', ' '),
                    category="Unknown",
                    icon_filename=icon_filename if icon_filename else None,
                    fallback_color=fallback_color,
                    description=f"System type: {name}",
                    sort_order=0
                )
                
                if success:
                    QMessageBox.information(dialog, "Success", f"System type '{name}' added successfully!")
                    dialog.close()
                    self.system_type_manager.load_system_types()
                else:
                    QMessageBox.critical(dialog, "Error", "Failed to add system type. It may already exist.")
                    
            except Exception as e:
                self.logger.error(f"Error adding system type: {e}")
                QMessageBox.critical(dialog, "Error", f"Failed to add system type: {str(e)}")
        
        icon_browse_btn.clicked.connect(browse_icon_file)
        add_button.clicked.connect(add_system_type)
        cancel_button.clicked.connect(dialog.close)
        
        dialog.show()

    def __del__(self):
        self.logger.info("MainApp destructor called, cleaning up resources...")
        self.close_all_windows()
    
    def handle_more_button(self):
        pass
    
    def setup_more_button_menu(self):
        more_button = getattr(self.window.ui, 'more_button', None)
        if more_button and hasattr(more_button, 'menu') and more_button.menu():
            menu = more_button.menu()
            
            msportals_action = QAction("Azure Portals", self.window)
            msportals_action.triggered.connect(self.display_msportals_data)
            menu.addAction(msportals_action)
            
            for action in menu.actions():
                if action.text() == "Windows - Event ID":
                    action.triggered.connect(self.display_event_id_kb)
                elif action.text() == "Windows - LOLBAS":
                    action.triggered.connect(self.display_lolbas_kb)
    
    def display_msportals_data(self):
        try:
            window = display_msportals_data(self, self.db_path)
            self.track_child_window(window)
        except Exception as e:
            self.logger.error(f"Error opening msportals data: {e}")
            QMessageBox.critical(self.window, "Error", f"Failed to open msportals data: {e}")
        
    def close_all_windows(self):
        if hasattr(self, 'child_windows'):
            for window in self.child_windows[:]:  
                try:
                    if (window and 
                        hasattr(window, 'isVisible') and 
                        isValid(window) and  
                        window.isVisible()):
                        window.close()
                except RuntimeError as e:
                    if "already deleted" in str(e):
                        self.logger.info(f"Window already deleted: {e}")
                    else:
                        self.logger.error(f"Error closing window: {e}")
                except Exception as e:
                    self.logger.error(f"Error closing window: {e}")
            self.child_windows.clear()

    def exit_application(self):
        self.close_all_windows()
        self.window.close()
        self.app.quit()

    def track_child_window(self, window):
        if window:
            if hasattr(window, 'setParent') and window.parent() is None:
                window.setParent(self.window, Qt.Window)
            self.child_windows.append(window)
            window.setAttribute(Qt.WA_DeleteOnClose, False)
            original_close = window.closeEvent if hasattr(window, 'closeEvent') else None
            
            def new_close_event(event):
                if window in self.child_windows:
                    self.child_windows.remove(window)
                if original_close:
                    original_close(event)
                else:
                    event.accept()
            window.closeEvent = new_close_event
            return window
        return None

    def application_cleanup(self):
        self.release_file_lock()
        if self.mitre_flow_window:
            try:
                self.mitre_flow_window.close()
                self.mitre_flow_window = None
                self.logger.info("Preloaded MITRE Flow window cleaned up")
            except Exception as e:
                self.logger.error(f"Error cleaning up preloaded MITRE Flow window: {e}")
        for window in self.child_windows[:]:
            try:
                if window and isValid(window):  
                    self.logger.info(f"Final cleanup of window: {window.windowTitle() if hasattr(window, 'windowTitle') else 'Unknown'}")
                    window.setAttribute(Qt.WA_DeleteOnClose, True)
                    window.close()
            except RuntimeError as e:
                if "already deleted" in str(e):
                    self.logger.info(f"Window already deleted: {e}")
                else:
                    self.logger.error(f"Error in final cleanup: {e}")
            except Exception as e:
                self.logger.error(f"Error in final cleanup: {e}")


if __name__ == "__main__":
    main_app = MainApp()
    sys.exit(main_app.run())